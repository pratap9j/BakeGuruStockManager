import os
import json
import sqlite3
from datetime import datetime
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from fpdf import FPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# -----------------------------
# App config
# -----------------------------
st.set_page_config(page_title="BakeGuru Stock Manager", layout="wide")

DB_FILE = "bakeguru.db"
IMAGES_DIR = "images"
os.makedirs(IMAGES_DIR, exist_ok=True)

# -----------------------------
# DB init (safe migrations)
# -----------------------------
def init_db():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT UNIQUE,
            name TEXT,
            category TEXT,
            subcategory TEXT,
            price REAL,
            stock INTEGER,
            image_url TEXT,
            image_path TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_no TEXT,
            date TEXT,
            customer_name TEXT,
            customer_company TEXT,
            customer_address TEXT,
            customer_phone TEXT,
            items TEXT,
            total REAL
        )
    """)

    # Migrate missing columns if DB is older
    def ensure_col(table, col, coltype):
        c.execute(f"PRAGMA table_info({table})")
        cols = [r[1] for r in c.fetchall()]
        if col not in cols:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coltype}")
    ensure_col("products", "image_path", "TEXT")

    conn.commit()
    return conn

conn = init_db()

# -----------------------------
# Header
# -----------------------------
col1, col2 = st.columns([1, 6])
with col1:
    st.markdown("### 📦")
with col2:
    st.markdown(
        "<h1 style='margin-bottom:0;'>BakeGuru Stock Manager</h1>"
        "<div style='color:#888;margin-top:2px;'>Inventory • Quotes • Dashboard</div>",
        unsafe_allow_html=True,
    )
st.markdown("---")

# -----------------------------
# Sidebar (list navigation)
# -----------------------------
st.sidebar.markdown("### 📂 Navigation")
MENU = ["Dashboard", "View Stock", "Add Stock", "Quote Builder", "Quotes History"]
choice = st.sidebar.radio("Go to", MENU, index=0)

# -----------------------------
# Utils
# -----------------------------
def load_products() -> pd.DataFrame:
    return pd.read_sql("SELECT * FROM products", conn)

def safe_items_json(df: pd.DataFrame) -> str:
    # Avoid ujson overflow / encoding issues
    return json.dumps(df.to_dict(orient="records"), ensure_ascii=False)

def local_image_bytes(path: str):
    if path and os.path.exists(path):
        try:
            with open(path, "rb") as f:
                return f.read()
        except Exception:
            return None
    return None

def pick_thumb(row):
    """
    Prefer local image_path (bytes) so uploaded images show first.
    Fallback to image_url; else None.
    """
    path = (row.get("image_path") or "").strip()
    if path and os.path.exists(path):
        b = local_image_bytes(path)
        if b:
            return b
    url = (row.get("image_url") or "").strip()
    return url if url else None

# -----------------------------
# Exports (PDF & Excel)
# -----------------------------
def export_quote_pdf(df, qno, name, company, addr, phone, total):
    pdf = FPDF()
    pdf.add_page()

    # Title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, "BakeGuru Quote", ln=True, align="C")
    pdf.ln(6)

    # Customer info
    pdf.set_font("Arial", "", 12)
    pdf.cell(200, 8, f"Quote No: {qno}", ln=True)
    pdf.cell(200, 8, f"Customer: {name} | {company}", ln=True)
    pdf.cell(200, 8, f"Phone: {phone}", ln=True)
    pdf.multi_cell(200, 8, f"Address: {addr}")
    pdf.ln(3)

    # Table header
    pdf.set_font("Arial", "B", 10)
    col_w = {"sku": 30, "name": 50, "qty": 18, "price": 25, "total": 25, "image": 40}
    pdf.cell(col_w["sku"], 8, "SKU", 1, align="C")
    pdf.cell(col_w["name"], 8, "Name", 1, align="C")
    pdf.cell(col_w["qty"], 8, "Qty", 1, align="C")
    pdf.cell(col_w["price"], 8, "Price", 1, align="C")
    pdf.cell(col_w["total"], 8, "Total", 1, align="C")
    pdf.cell(col_w["image"], 8, "Image", 1, ln=True, align="C")

    pdf.set_font("Arial", "", 10)

    for _, row in df.iterrows():
        qty = int(row["qty"])
        line_total = qty * float(row["price"])
        line_h = 22

        pdf.cell(col_w["sku"], line_h, str(row["sku"]), 1)
        pdf.cell(col_w["name"], line_h, str(row["name"])[:40], 1)
        pdf.cell(col_w["qty"], line_h, str(qty), 1, align="C")
        pdf.cell(col_w["price"], line_h, f"{float(row['price']):.2f}", 1, align="R")
        pdf.cell(col_w["total"], line_h, f"{line_total:.2f}", 1, align="R")

        # image cell
        inserted = False
        url = (row.get("image_url") or "").strip()
        path = (row.get("image_path") or "").strip()
        pdf.cell(col_w["image"], line_h, "", 1)
        x_left = pdf.get_x() - col_w["image"]
        y_top = pdf.get_y()
        try:
            if url:
                resp = requests.get(url, timeout=5)
                img = Image.open(BytesIO(resp.content)).convert("RGB")
                img.thumbnail((30, 30))
                tf = f"__tmp_{row['id']}.jpg"
                img.save(tf, "JPEG")
                pdf.image(tf, x=x_left + 5, y=y_top + 3, w=30, h=16)
                os.remove(tf)
                inserted = True
            elif path and os.path.exists(path):
                pdf.image(path, x=x_left + 5, y=y_top + 3, w=30, h=16)
                inserted = True
        except Exception:
            inserted = False
        pdf.ln(line_h)

    # Grand Total under the "Total" column (further left)
    x_total_col = pdf.l_margin + col_w["sku"] + col_w["name"] + col_w["qty"] + col_w["price"]
    pdf.ln(2)
    pdf.set_font("Arial", "B", 12)
    pdf.set_x(x_total_col)
    pdf.cell(col_w["total"], 10, "Grand Total:", border=0, align="R")
    pdf.cell(col_w["image"], 10, f"{total:.2f}", border=0, align="R")
    pdf.ln(10)

    # Disclaimer
    pdf.set_font("Arial", "I", 10)
    pdf.multi_cell(200, 7, "Note: GST & Shipping extra.")

    os.makedirs("exports", exist_ok=True)
    file_path = f"exports/{qno}.pdf"
    pdf.output(file_path)
    return file_path

def export_quote_excel(df, qno, name, company, addr, phone, total):
    os.makedirs("exports", exist_ok=True)
    file_path = f"exports/{qno}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    headers = ["SKU", "Name", "Qty", "Price", "Total", "Image"]
    ws.append(headers)

    for idx, row in df.iterrows():
        excel_row = idx + 2
        ws.append([
            row["sku"],
            row["name"],
            int(row["qty"]),
            float(row["price"]),
            float(row["qty"] * row["price"]),
            ""
        ])

        # Try URL then local path
        url = (row.get("image_url") or "").strip()
        path = (row.get("image_path") or "").strip()
        try:
            if url:
                resp = requests.get(url, timeout=5)
                img = Image.open(BytesIO(resp.content)).convert("RGB")
                img.thumbnail((80, 80))
                tf = f"__tmp_xl_{row['id']}.jpg"
                img.save(tf, "JPEG")
                xl_img = XLImage(tf)
                xl_img.width, xl_img.height = 60, 60
                ws.add_image(xl_img, f"F{excel_row}")
                os.remove(tf)
            elif path and os.path.exists(path):
                xl_img = XLImage(path)
                xl_img.width, xl_img.height = 60, 60
                ws.add_image(xl_img, f"F{excel_row}")
        except Exception:
            pass

    # widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Customer & total
    base = len(df) + 3
    ws[f"A{base}"] = "Customer:"
    ws[f"B{base}"] = name
    ws[f"A{base+1}"] = "Company:"
    ws[f"B{base+1}"] = company
    ws[f"A{base+2}"] = "Address:"
    ws[f"B{base+2}"] = addr
    ws[f"A{base+3}"] = "Phone:"
    ws[f"B{base+3}"] = phone
    ws[f"A{base+5}"] = "Grand Total"
    ws[f"B{base+5}"] = float(total)

    wb.save(file_path)
    return file_path

# -----------------------------
# Pages
# -----------------------------
def page_dashboard():
    st.subheader("📊 Dashboard")

    inv = load_products()
    if inv.empty:
        st.info("No inventory yet.")
    else:
        total_skus = len(inv)
        total_in_stock = int(inv["stock"].sum())
        out_of_stock = int((inv["stock"] == 0).sum())

        c1, c2, c3 = st.columns(3)
        c1.metric("Total SKUs", total_skus)
        c2.metric("Total Units In Stock", total_in_stock)
        c3.metric("Out of Stock SKUs", out_of_stock)

        st.write("### SKUs per Category")
        cat_counts = inv["category"].fillna("Uncategorized").value_counts()
        if not cat_counts.empty:
            st.bar_chart(cat_counts)

    q = pd.read_sql("SELECT * FROM quotes", conn)
    if not q.empty:
        st.write("### Quotes per Month")
        q["date"] = pd.to_datetime(q["date"], errors="coerce")
        monthly = q.groupby(q["date"].dt.to_period("M")).size()
        st.line_chart(monthly)

def page_view_stock():
    st.subheader("📦 View Stock")

    df = load_products()
    if df.empty:
        st.info("No products available.")
        return

    # Filters
    cols = st.columns([3, 2, 2, 2])
    with cols[0]:
        term = st.text_input("🔎 Search (SKU / Name / Category)")
    with cols[1]:
        cat = st.selectbox("Category", ["All"] + sorted(df["category"].dropna().unique().tolist()))
    with cols[2]:
        subcat = st.selectbox("Subcategory", ["All"] + sorted(df["subcategory"].dropna().unique().tolist()))
    with cols[3]:
        low = st.checkbox("Low stock (< 5)")

    f = df.copy()
    if term:
        mask = f.apply(lambda r: r.astype(str).str.contains(term, case=False, na=False).any(), axis=1)
        f = f[mask]
    if cat != "All":
        f = f[f["category"] == cat]
    if subcat != "All":
        f = f[f["subcategory"] == subcat]
    if low:
        f = f[f["stock"] < 5]

    # Thumbnail column
    thumbs = f.apply(pick_thumb, axis=1)
    f = f.copy()
    f.insert(1, "thumb", thumbs)

    st.markdown("### Products (inline editable)")
    edited = st.data_editor(
        f[["thumb", "sku", "name", "category", "subcategory", "price", "stock", "image_url", "image_path", "id"]],
        column_config={
            "thumb": st.column_config.ImageColumn("Image", help="Local image (image_path) preferred, else URL"),
            "sku": st.column_config.TextColumn("SKU", disabled=True),
            "name": "Name",
            "category": "Category",
            "subcategory": "Subcategory",
            "price": st.column_config.NumberColumn("Price", step=0.01, format="%.2f"),
            "stock": st.column_config.NumberColumn("Stock", step=1),
            "image_url": st.column_config.TextColumn("Image URL"),
            "image_path": st.column_config.TextColumn("Local Image Path", disabled=True),
            "id": st.column_config.TextColumn("ID", disabled=True),
        },
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        height=420,
        key="inv_editor",
    )

    if st.button("💾 Save changes"):
        try:
            for _, r in edited.iterrows():
                conn.execute(
                    "UPDATE products SET name=?, category=?, subcategory=?, price=?, stock=?, image_url=? WHERE id=?",
                    (str(r["name"]), str(r["category"]), str(r["subcategory"]),
                     float(r["price"]), int(r["stock"]), str(r["image_url"] or ""), int(r["id"]))
                )
            conn.commit()
            st.success("Changes saved.")
            st.rerun()
        except Exception as e:
            st.error(f"Save failed: {e}")

    # Image manager (set/replace/clear image_path)
    st.markdown("### 🖼 Manage product image (image_path)")
    left, right = st.columns([2, 3])
    with left:
        df_all = load_products()
        if df_all.empty:
            st.info("No products to manage.")
            return
        choices = [f"{r['sku']} — {r['name']}" for _, r in df_all.iterrows()]
        choice_map = {f"{r['sku']} — {r['name']}": int(r["id"]) for _, r in df_all.iterrows()}
        pick = st.selectbox("Pick a product", choices)

    with right:
        up = st.file_uploader("Upload JPG/PNG to set/replace image_path", type=["jpg", "jpeg", "png"],
                              key="img_uploader_viewstock")

    c1, c2, _ = st.columns([1, 1, 3])

    with c1:
        if st.button("Upload & Set"):
            if not pick:
                st.warning("Select a product first.")
            elif up is None:
                st.warning("Choose a file to upload.")
            else:
                try:
                    pid = choice_map[pick]
                    cur = conn.execute("SELECT sku, image_path FROM products WHERE id=?", (pid,))
                    sku, old_path = cur.fetchone()
                    safe_sku = "".join(ch for ch in sku if ch.isalnum() or ch in ("-", "_")).strip()
                    ext = os.path.splitext(up.name)[1].lower() or ".jpg"
                    fname = f"{safe_sku}_{int(datetime.now().timestamp())}{ext}"
                    new_path = os.path.join(IMAGES_DIR, fname)

                    img = Image.open(up).convert("RGB")
                    img.save(new_path, "JPEG", quality=90)

                    # delete old local image if existed
                    try:
                        if old_path and os.path.exists(old_path):
                            os.remove(old_path)
                    except Exception:
                        pass

                    conn.execute("UPDATE products SET image_path=? WHERE id=?", (new_path, pid))
                    conn.commit()
                    st.success(f"Image set for {pick}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to set image: {e}")

    with c2:
        if st.button("Clear image"):
            if not pick:
                st.warning("Select a product first.")
            else:
                try:
                    pid = choice_map[pick]
                    cur = conn.execute("SELECT image_path FROM products WHERE id=?", (pid,))
                    (old_path,) = cur.fetchone()
                    try:
                        if old_path and os.path.exists(old_path):
                            os.remove(old_path)
                    except Exception:
                        pass
                    conn.execute("UPDATE products SET image_path='' WHERE id=?", (pid,))
                    conn.commit()
                    st.success(f"Cleared local image for {pick}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to clear image: {e}")

    st.caption("Tip: Local image (image_path) is shown in thumbnails and exports when present; otherwise Image URL is used.")

def page_add_stock():
    st.subheader("➕ Add Stock")

    with st.form("add_stock_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            sku = st.text_input("SKU")
            name = st.text_input("Product Name")
            category = st.text_input("Category")
            subcategory = st.text_input("Subcategory")
        with c2:
            price = st.number_input("Price", min_value=0.0, step=0.01, format="%.2f")
            stock = st.number_input("Stock Quantity", min_value=0, step=1)
            image_url = st.text_input("Image URL (optional)")

        img_file = st.file_uploader("Or upload an image file (JPG/PNG)", type=["jpg", "jpeg", "png"])

        submitted = st.form_submit_button("Add Product")

        if submitted:
            if not sku or not name:
                st.error("Please provide at least SKU and Product Name.")
                return

            image_path = ""
            if img_file is not None:
                try:
                    ext = os.path.splitext(img_file.name)[1].lower() or ".jpg"
                    safe_sku = "".join(ch for ch in sku if ch.isalnum() or ch in ("-", "_")).strip()
                    fname = f"{safe_sku}_{int(datetime.now().timestamp())}{ext}"
                    image_path = os.path.join(IMAGES_DIR, fname)
                    img = Image.open(img_file).convert("RGB")
                    img.save(image_path, "JPEG", quality=90)
                except Exception as e:
                    st.warning(f"Image save failed (continuing without local image): {e}")

            try:
                conn.execute(
                    "INSERT INTO products (sku, name, category, subcategory, price, stock, image_url, image_path) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (sku, name, category, subcategory, float(price), int(stock), image_url or "", image_path or "")
                )
                conn.commit()
                st.success(f"✅ {name} added.")
            except sqlite3.IntegrityError:
                st.error("SKU already exists. Try another SKU.")

    st.markdown("#### 📤 Bulk Upload (Excel .xlsx)")
    st.caption("Columns required: sku, name, category, subcategory, price, stock, image_url (optional).")
    up = st.file_uploader("Upload file", type=["xlsx"])
    if up:
        try:
            data = pd.read_excel(up)
            data.columns = [c.strip().lower() for c in data.columns]
            req = ["sku", "name", "category", "subcategory", "price", "stock"]
            missing = [c for c in req if c not in data.columns]
            if missing:
                st.error(f"Missing columns: {', '.join(missing)}")
            else:
                ok = 0
                for _, r in data.iterrows():
                    try:
                        conn.execute(
                            "INSERT INTO products (sku, name, category, subcategory, price, stock, image_url, image_path) "
                            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                            (str(r["sku"]), str(r["name"]), str(r.get("category") or ""), str(r.get("subcategory") or ""),
                             float(r.get("price") or 0), int(r.get("stock") or 0), str(r.get("image_url") or ""), "")
                        )
                        ok += 1
                    except sqlite3.IntegrityError:
                        pass
                conn.commit()
                st.success(f"✅ Bulk upload complete. Inserted {ok} rows.")
        except Exception as e:
            st.error(f"Upload failed: {e}")

def page_quote_builder():
    st.subheader("🧾 Quote Builder")

    df = load_products()
    if df.empty:
        st.info("No products in inventory.")
        return

    if "quote_cart" not in st.session_state:
        st.session_state.quote_cart = []

    # Search/filter
    s1, _ = st.columns([3, 1])
    with s1:
        q = st.text_input("Search products (SKU/Name/Category)")

    filt = df.copy()
    if q:
        filt = filt[filt.apply(lambda r: r.astype(str).str.contains(q, case=False, na=False).any(), axis=1)]

    st.write("### Add Items")
    for _, row in filt.iterrows():
        c1, c2, c3, c4, c5 = st.columns([1.2, 3, 1.2, 1.2, 1.2])
        with c1:
            thumb = pick_thumb(row)
            try:
                if isinstance(thumb, (bytes, bytearray)) or (isinstance(thumb, str) and thumb):
                    st.image(thumb, width=50)
                else:
                    st.write("📦")
            except Exception:
                st.write("📦")
        with c2:
            st.write(f"**{row['name']}**")
            st.caption(f"SKU: {row['sku']} • {row['category']}/{row['subcategory']}")
        with c3:
            st.write(f"₹{row['price']:.2f}")
        with c4:
            qty = st.number_input(
                f"Qty_{row['id']}",
                min_value=1,
                max_value=int(row['stock']) if row['stock'] else 9999,
                value=1, step=1, key=f"qty_{row['id']}"
            )
        with c5:
            if st.button("Add to Quote", key=f"add_{row['id']}"):
                cart = st.session_state.quote_cart
                existing = next((i for i, it in enumerate(cart) if it["id"] == row["id"]), None)
                item = {**row.to_dict(), "qty": int(qty)}
                if existing is None:
                    cart.append(item)
                else:
                    cart[existing]["qty"] += int(qty)
                st.success(f"Added {row['name']} x{qty}")

    st.markdown("### 🛒 Current Quote Cart")
    if not st.session_state.quote_cart:
        st.info("No items added yet.")
        return

    cart_df = pd.DataFrame(st.session_state.quote_cart)
    cart_df["line_total"] = cart_df["qty"] * cart_df["price"]
    st.dataframe(cart_df[["sku", "name", "qty", "price", "line_total"]], use_container_width=True)
    grand_total = float(cart_df["line_total"].sum())
    st.markdown(f"**Grand Total:** ₹ {grand_total:,.2f}")

    st.markdown("---")
    st.markdown("### 👤 Customer Details")
    c1, c2 = st.columns(2)
    with c1:
        cust_name = st.text_input("Name")
        cust_company = st.text_input("Company")
    with c2:
        cust_phone = st.text_input("Phone")
        cust_addr = st.text_area("Address", height=92)

    cta1, cta2, _ = st.columns([1, 1, 4])
    with cta1:
        if st.button("🧾 Generate PDF"):
            today = datetime.today().strftime("%Y-%m-%d")
            qcount = pd.read_sql("SELECT COUNT(*) AS c FROM quotes", conn)["c"][0] + 1
            qno = f"BG-{qcount:04d}"
            items_df = pd.DataFrame(st.session_state.quote_cart)
            total = float((items_df["qty"] * items_df["price"]).sum())

            conn.execute(
                "INSERT INTO quotes (quote_no, date, customer_name, customer_company, customer_address, customer_phone, items, total) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (qno, today, cust_name, cust_company, cust_addr, cust_phone, safe_items_json(items_df), total),
            )
            conn.commit()

            pdf_file = export_quote_pdf(items_df, qno, cust_name, cust_company, cust_addr, cust_phone, total)
            st.success(f"✅ Quote {qno} saved.")
            with open(pdf_file, "rb") as f:
                st.download_button("📥 Download PDF", f, file_name=f"{qno}.pdf")

    with cta2:
        if st.button("📊 Export Excel"):
            qcount = pd.read_sql("SELECT COUNT(*) AS c FROM quotes", conn)["c"][0] + 1
            qno = f"BG-{qcount:04d}"
            items_df = pd.DataFrame(st.session_state.quote_cart)
            total = float((items_df["qty"] * items_df["price"]).sum())
            xls = export_quote_excel(items_df, qno, cust_name, cust_company, cust_addr, cust_phone, total)
            with open(xls, "rb") as f:
                st.download_button("⬇️ Download Excel", f, file_name=f"{qno}.xlsx")

    st.markdown("#### 🗑 Manage Cart")
    rm_skus = st.multiselect(
        "Remove items from cart",
        [f"{r['sku']} - {r['name']}" for r in st.session_state.quote_cart]
    )
    if st.button("Remove Selected"):
        if rm_skus:
            st.session_state.quote_cart = [
                r for r in st.session_state.quote_cart
                if f"{r['sku']} - {r['name']}" not in rm_skus
            ]
            st.success("Removed selected items.")

def page_quotes_history():
    st.subheader("📜 Quotes History")
    q = pd.read_sql("SELECT id, quote_no, date, customer_name, total FROM quotes ORDER BY id DESC", conn)
    if q.empty:
        st.info("No quotes yet.")
        return
    st.dataframe(q, use_container_width=True)

# -----------------------------
# Router
# -----------------------------
if choice == "Dashboard":
    page_dashboard()
elif choice == "View Stock":
    page_view_stock()
elif choice == "Add Stock":
    page_add_stock()
elif choice == "Quote Builder":
    page_quote_builder()
elif choice == "Quotes History":
    page_quotes_history()
